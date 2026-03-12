"""
dashboard/banking/rapport.py
Génération des rapports — PDF et Excel — par banque (données MongoDB)

FONCTIONNALITÉS :
   Rapport PDF complet professionnel (6 sections)
   Export Excel multi-feuilles (tableau de bord, historique, ratios, positionnement)
   Gestion gracieuse des données manquantes
   Ratios recalculés si absents
   Évaluation des ratios vs normes BCEAO
"""
import io
import base64
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable,
)

# ── Couleurs charte bancaire ──────────────────────────────────────────────────
C_BLEU    = colors.HexColor("#1A3C6F")
C_BLEU_L  = colors.HexColor("#2563EB")
C_FONCE   = colors.HexColor("#0F172A")
C_VERT    = colors.HexColor("#059669")
C_ORANGE  = colors.HexColor("#D97706")
C_GRIS    = colors.HexColor("#64748B")
C_FOND    = colors.HexColor("#F8FAFC")
C_FOND2   = colors.HexColor("#EFF6FF")
C_BLANC   = colors.white
C_ALERTE  = colors.HexColor("#FEF3C7")
C_ALERTE_B = colors.HexColor("#D97706")
C_EMERALD = colors.HexColor("#10B981")


def _fmt(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "N/D"
    v = float(val)
    if abs(v) >= 1_000_000:
        return f"{v/1_000_000:.2f} B FCFA"
    if abs(v) >= 1_000:
        return f"{v/1_000:.1f} Mds FCFA"
    return f"{v:,.0f} M FCFA"


def _pct(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "N/D"
    return f"{float(val):.2f} %"


def _recalcul_ratios(df):
    """Recalcule les ratios manquants dans un DataFrame."""
    if "ratio_solvabilite" not in df.columns or df["ratio_solvabilite"].isna().all():
        if "fonds_propres" in df.columns and "bilan" in df.columns:
            df["ratio_solvabilite"] = (df["fonds_propres"] / df["bilan"] * 100).round(4)
    if "ratio_rendement_actifs" not in df.columns or df["ratio_rendement_actifs"].isna().all():
        if "resultat_net" in df.columns and "bilan" in df.columns:
            df["ratio_rendement_actifs"] = (df["resultat_net"] / df["bilan"] * 100).round(4)
    if "ratio_rentabilite_capitaux" not in df.columns or df["ratio_rentabilite_capitaux"].isna().all():
        if "resultat_net" in df.columns and "fonds_propres" in df.columns:
            df["ratio_rentabilite_capitaux"] = (df["resultat_net"] / df["fonds_propres"] * 100).round(4)
    if "coefficient_exploitation" not in df.columns or df["coefficient_exploitation"].isna().all():
        if "charges_generales_exploitation" in df.columns and "produit_net_bancaire" in df.columns:
            df["coefficient_exploitation"] = (df["charges_generales_exploitation"] / df["produit_net_bancaire"] * 100).round(4)
    if "ratio_emplois_ressources" not in df.columns or df["ratio_emplois_ressources"].isna().all():
        if "emploi" in df.columns and "ressources" in df.columns:
            df["ratio_emplois_ressources"] = (df["emploi"] / df["ressources"].replace(0, float("nan")) * 100).round(4)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# RAPPORT PDF
# ══════════════════════════════════════════════════════════════════════════════
def generer_rapport_pdf(banque: str, annee: int) -> str:
    """
    Génère un rapport PDF complet pour une banque.
    Retourne le contenu PDF en base64 (pour dcc.Download).
    """
    from utils.db import get_collection

    col, client = get_collection()
    docs = list(col.find({"sigle": banque}, {"_id": 0}))
    client.close()

    if not docs:
        raise ValueError(f"Aucune donnée dans MongoDB pour la banque : {banque}")

    df = pd.DataFrame(docs)
    num_cols = ["bilan", "fonds_propres", "resultat_net", "produit_net_bancaire",
                "charges_generales_exploitation", "emploi", "ressources",
                "cout_du_risque", "impots_benefices", "interets_produits",
                "interets_charges", "commissions_produits", "commissions_charges"]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    df = _recalcul_ratios(df)

    row_df = df[df["annee"] == annee]
    annee_manquante = row_df.empty

    if annee_manquante:
        annees_dispo = sorted(df["annee"].dropna().unique().astype(int).tolist())
        annee_ref = annees_dispo[-1]
        row = df[df["annee"] == annee_ref].iloc[0]
    else:
        annee_ref = annee
        row = row_df.iloc[0]

    col2, client2 = get_collection()
    docs_all = list(col2.find({"annee": annee_ref}, {"_id": 0}))
    client2.close()
    df_all = pd.DataFrame(docs_all) if docs_all else pd.DataFrame()
    if not df_all.empty:
        for c in num_cols:
            if c in df_all.columns:
                df_all[c] = pd.to_numeric(df_all[c], errors="coerce")

    # ── Document ──────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=1.8*cm, bottomMargin=2*cm,
        title=f"Rapport Bancaire {banque} {annee}",
    )
    story = []

    def s_titre(txt, size=18, color=C_BLANC):
        return Paragraph(txt, ParagraphStyle(
            "t", fontSize=size, fontName="Helvetica-Bold", textColor=color, spaceAfter=2))

    def s_section(txt, color=C_BLEU):
        return Paragraph(txt, ParagraphStyle(
            "s", fontSize=11, fontName="Helvetica-Bold",
            textColor=color, spaceBefore=10, spaceAfter=6))

    def s_body(txt):
        return Paragraph(txt, ParagraphStyle(
            "b", fontSize=8.5, fontName="Helvetica",
            textColor=colors.HexColor("#374151"), leading=14))

    def s_small(txt, color=C_GRIS):
        return Paragraph(txt, ParagraphStyle(
            "sm", fontSize=7.5, fontName="Helvetica", textColor=color, leading=11))

    def s_alerte(txt):
        return Paragraph(txt, ParagraphStyle(
            "al", fontSize=8.5, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#92400E"), leading=14))

    # EN-TÊTE
    entete_data = [[s_titre("RAPPORT DE POSITIONNEMENT BANCAIRE", 14), s_titre(banque, 28, C_EMERALD)]]
    entete = Table(entete_data, colWidths=["60%", "40%"])
    entete.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), C_FONCE),
        ("TOPPADDING",    (0,0),(-1,-1), 22),
        ("BOTTOMPADDING", (0,0),(-1,-1), 18),
        ("LEFTPADDING",   (0,0),(-1,-1), 22),
        ("RIGHTPADDING",  (0,0),(-1,-1), 22),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
    ]))
    story.append(entete)
    story.append(Spacer(1, 0.3*cm))

    accent_data = [[s_small(
        f"Groupe : <b>{row.get('groupe_bancaire','N/D')}</b>  ·  "
        f"Année de référence : <b>{annee_ref}</b>  ·  "
        f"Source : BCEAO — MongoDB Atlas  ·  DataSphere M2 Big Data ISM",
        color=C_BLANC
    )]]
    t_accent = Table(accent_data, colWidths=["100%"])
    t_accent.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), C_BLEU),
        ("TOPPADDING",    (0,0),(-1,-1), 6),
        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
        ("LEFTPADDING",   (0,0),(-1,-1), 14),
    ]))
    story.append(t_accent)
    story.append(Spacer(1, 0.3*cm))

    if annee_manquante:
        t_alerte = Table([[s_alerte(
            f"⚠️  Données non disponibles pour {banque} en {annee}. "
            f"Ce rapport utilise les données de {annee_ref} (dernière année disponible)."
        )]], colWidths=["100%"])
        t_alerte.setStyle(TableStyle([
            ("BACKGROUND", (0,0),(-1,-1), C_ALERTE),
            ("LINEBEFORE",  (0,0),(0,-1), 4, C_ALERTE_B),
            ("TOPPADDING", (0,0),(-1,-1), 8),
            ("BOTTOMPADDING", (0,0),(-1,-1), 8),
            ("LEFTPADDING", (0,0),(-1,-1), 14),
        ]))
        story.append(t_alerte)
        story.append(Spacer(1, 0.3*cm))

    # SYNTHÈSE SECTORIELLE
    story.append(s_section(f"1. Synthèse du Secteur Bancaire Sénégalais — {annee_ref}"))
    nb_banques  = df_all["sigle"].nunique() if not df_all.empty else "N/D"
    total_bilan = df_all["bilan"].sum()     if "bilan" in df_all.columns else None
    total_pnb   = df_all["produit_net_bancaire"].sum() if "produit_net_bancaire" in df_all.columns else None
    story.append(s_body(
        f"Le secteur bancaire sénégalais compte <b>{nb_banques} banques</b> actives en {annee_ref}, "
        f"représentant un total bilan de <b>{_fmt(total_bilan)}</b> "
        f"et un PNB global de <b>{_fmt(total_pnb)}</b>. "
        f"La banque <b>{banque}</b> appartient au groupe <b>{row.get('groupe_bancaire','N/D')}</b>."
    ))
    story.append(Spacer(1, 0.3*cm))

    # KPIs
    story.append(s_section(f"2. Indicateurs Clés — {banque} ({annee_ref})"))
    kpi_data = [
        ["INDICATEUR", "VALEUR", "INDICATEUR", "VALEUR"],
        ["Total Bilan",          _fmt(row.get("bilan")),         "Produit Net Bancaire", _fmt(row.get("produit_net_bancaire"))],
        ["Emploi (Crédits)",     _fmt(row.get("emploi")),        "Ressources (Dépôts)",  _fmt(row.get("ressources"))],
        ["Fonds Propres",        _fmt(row.get("fonds_propres")), "Résultat Net",         _fmt(row.get("resultat_net"))],
        ["Intérêts Produits",    _fmt(row.get("interets_produits")), "Intérêts Charges", _fmt(row.get("interets_charges"))],
        ["Commissions Produits", _fmt(row.get("commissions_produits")), "Commissions Charges", _fmt(row.get("commissions_charges"))],
        ["Coût du Risque",       _fmt(row.get("cout_du_risque")), "Impôts / Bénéfices",  _fmt(row.get("impots_benefices"))],
    ]
    t_kpi = Table(kpi_data, colWidths=["30%", "20%", "30%", "20%"])
    t_kpi.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1, 0), C_BLEU_L),
        ("TEXTCOLOR",     (0,0),(-1, 0), C_BLANC),
        ("FONTNAME",      (0,0),(-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_BLANC, C_FOND2]),
        ("ALIGN",         (1,0),(1,-1),  "RIGHT"),
        ("ALIGN",         (3,0),(3,-1),  "RIGHT"),
        ("GRID",          (0,0),(-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",    (0,0),(-1,-1), 6),
        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
    ]))
    story.append(t_kpi)
    story.append(Spacer(1, 0.35*cm))

    # RATIOS FINANCIERS
    story.append(s_section("3. Analyse des Ratios Financiers"))

    def _eval_ratio(nom, val, seuil_ok, seuil_warn, plus_haut=True):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "N/D"
        v = float(val)
        if plus_haut:
            if v >= seuil_ok:   return f"✅ Bon ({v:.2f}%)"
            if v >= seuil_warn: return f"⚠️ Moyen ({v:.2f}%)"
            return f"❌ Insuffisant ({v:.2f}%)"
        else:
            if v <= seuil_ok:   return f"✅ Bon ({v:.2f}%)"
            if v <= seuil_warn: return f"⚠️ Moyen ({v:.2f}%)"
            return f"❌ Élevé ({v:.2f}%)"

    ratio_data = [
        ["RATIO", "VALEUR", "NORME BCEAO", "ÉVALUATION"],
        ["ROA — Rendement des actifs",         _pct(row.get("ratio_rendement_actifs")),     "Secteur ~1–2 %",      _eval_ratio("ROA", row.get("ratio_rendement_actifs"), 2, 1)],
        ["ROE — Rentabilité fonds propres",    _pct(row.get("ratio_rentabilite_capitaux")), "Secteur ~10–15 %",    _eval_ratio("ROE", row.get("ratio_rentabilite_capitaux"), 10, 5)],
        ["Ratio de solvabilité (FP/Bilan)",   _pct(row.get("ratio_solvabilite")),          "≥ 8 % (norme BCEAO)", _eval_ratio("Solv", row.get("ratio_solvabilite"), 12, 8)],
        ["Coeff. d'exploitation (Charges/PNB)",_pct(row.get("coefficient_exploitation")),  "< 65 % (optimal)",    _eval_ratio("Coeff", row.get("coefficient_exploitation"), 65, 80, False)],
        ["Ratio liquidité (Emploi/Ressources)",_pct(row.get("ratio_emplois_ressources")),  "< 100 % (prudentiel)",_eval_ratio("Liq", row.get("ratio_emplois_ressources"), 100, 110, False)],
    ]
    t_rat = Table(ratio_data, colWidths=["31%", "15%", "22%", "32%"])
    t_rat.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1, 0), C_VERT),
        ("TEXTCOLOR",     (0,0),(-1, 0), C_BLANC),
        ("FONTNAME",      (0,0),(-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_BLANC, C_FOND]),
        ("ALIGN",         (1,0),(2,-1),  "CENTER"),
        ("GRID",          (0,0),(-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",    (0,0),(-1,-1), 6),
        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
    ]))
    story.append(t_rat)
    story.append(Spacer(1, 0.35*cm))

    # ÉVOLUTION HISTORIQUE
    story.append(s_section(f"4. Évolution Historique — {banque}"))
    histo = df.sort_values("annee")[
        ["annee", "bilan", "produit_net_bancaire", "resultat_net", "fonds_propres", "emploi", "ressources"]
    ].dropna(subset=["bilan"])
    histo_rows = [["Année", "Bilan", "PNB", "Rés. Net", "Fonds Propres", "Emploi", "Ressources"]]
    for _, r in histo.iterrows():
        try:
            yr = int(r["annee"])
        except Exception:
            continue
        histo_rows.append([
            str(yr), _fmt(r.get("bilan")), _fmt(r.get("produit_net_bancaire")),
            _fmt(r.get("resultat_net")), _fmt(r.get("fonds_propres")),
            _fmt(r.get("emploi")), _fmt(r.get("ressources")),
        ])
    t_histo = Table(histo_rows, colWidths=["9%","16%","14%","14%","16%","14%","17%"])
    t_histo.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1, 0), C_FONCE),
        ("TEXTCOLOR",     (0,0),(-1, 0), C_BLANC),
        ("FONTNAME",      (0,0),(-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 7),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_BLANC, C_FOND]),
        ("ALIGN",         (1,0),(-1,-1), "RIGHT"),
        ("ALIGN",         (0,0),(0,-1),  "CENTER"),
        ("GRID",          (0,0),(-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
    ]))
    story.append(t_histo)
    story.append(Spacer(1, 0.35*cm))

    # POSITIONNEMENT CONCURRENTIEL
    story.append(s_section(f"5. Positionnement Concurrentiel — Bilan {annee_ref}"))
    if not df_all.empty and "bilan" in df_all.columns:
        classmt = (
            df_all[["sigle", "bilan"]].dropna()
            .sort_values("bilan", ascending=False)
            .reset_index(drop=True)
        )
        classmt.index += 1
        rang = classmt[classmt["sigle"] == banque].index
        rang_txt = f"#{rang[0]}" if len(rang) > 0 else "N/D"
        nb_tot = len(classmt)
        bilan_val = row.get("bilan")
        total_b = df_all["bilan"].sum()
        part = round(float(bilan_val) / float(total_b) * 100, 2) if (bilan_val and total_b) else None
        story.append(s_body(
            f"La banque <b>{banque}</b> se positionne <b>{rang_txt} sur {nb_tot}</b> banques "
            f"en termes de total bilan ({_fmt(bilan_val)}) en {annee_ref}."
            + (f" Part de marché bilan : <b>{part:.2f}%</b>." if part else "")
        ))
        story.append(Spacer(1, 0.2*cm))
        top8 = classmt.head(8)
        top8_data = [["Rang", "Banque", "Bilan (M FCFA)"]]
        for idx, r2 in top8.iterrows():
            marker = " ◀ (vous)" if r2["sigle"] == banque else ""
            top8_data.append([f"#{idx}", f"{r2['sigle']}{marker}", f"{r2['bilan']:,.0f}"])
        t_top = Table(top8_data, colWidths=["12%", "50%", "38%"])
        t_top.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1, 0), C_ORANGE),
            ("TEXTCOLOR",     (0,0),(-1, 0), C_BLANC),
            ("FONTNAME",      (0,0),(-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_BLANC, C_FOND]),
            ("ALIGN",         (2,0),(2,-1),  "RIGHT"),
            ("ALIGN",         (0,0),(0,-1),  "CENTER"),
            ("GRID",          (0,0),(-1,-1), 0.4, colors.HexColor("#E2E8F0")),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
            ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ]))
        story.append(t_top)

    # PIED DE PAGE
    story.append(Spacer(1, 0.6*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=C_GRIS))
    story.append(Spacer(1, 0.15*cm))
    note = (f"⚠ Données de {annee_ref} utilisées (données {annee} non disponibles). · " if annee_manquante else "")
    story.append(s_small(
        f"{note}Rapport généré automatiquement · DataSphere Sénégal · "
        f"Banque : {banque} · Source : BCEAO · MongoDB Atlas · M2 Big Data ISM · "
        "Les valeurs sont exprimées en millions de FCFA.",
        color=C_GRIS
    ))

    doc.build(story)
    return base64.b64encode(buffer.getvalue()).decode("utf-8")


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def generer_export_excel(banque: str, annee: int) -> bytes:
    """
    Génère un export Excel complet pour une banque.
    Retourne les bytes du fichier Excel.
    Nécessite : pip install openpyxl
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl requis : pip install openpyxl")

    from utils.db import get_collection

    col, client = get_collection()
    docs_banque = list(col.find({"sigle": banque}, {"_id": 0}))
    docs_all_annee = list(col.find({"annee": annee}, {"_id": 0}))
    client.close()

    if not docs_banque:
        raise ValueError(f"Aucune donnée pour : {banque}")

    df_banque = pd.DataFrame(docs_banque).sort_values("annee")
    df_all = pd.DataFrame(docs_all_annee) if docs_all_annee else pd.DataFrame()

    num_cols = ["bilan", "fonds_propres", "resultat_net", "produit_net_bancaire",
                "charges_generales_exploitation", "emploi", "ressources",
                "cout_du_risque", "interets_produits", "interets_charges"]
    for c in num_cols:
        if c in df_banque.columns:
            df_banque[c] = pd.to_numeric(df_banque[c], errors="coerce")
        if not df_all.empty and c in df_all.columns:
            df_all[c] = pd.to_numeric(df_all[c], errors="coerce")

    df_banque = _recalcul_ratios(df_banque)

    wb = openpyxl.Workbook()

    # Helpers de style
    def fill(hex_c):
        return PatternFill("solid", fgColor=hex_c.lstrip("#"))

    def font(bold=False, size=10, color="FFFFFF", name="Calibri"):
        return Font(name=name, bold=bold, size=size, color=color)

    def align(h="center", v="center", wrap=True):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def header_row(ws, row_i, headers, bg="1E3A5F"):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=row_i, column=j, value=h)
            c.fill = fill(bg)
            c.font = font(bold=True, size=10)
            c.alignment = align()
            c.border = border()

    def style_cell(c, value, bg="F8FAFC", fg="64748B", bold=False, h="center"):
        c.value = value
        c.fill = fill(bg)
        c.font = font(bold=bold, size=10, color=fg)
        c.alignment = align(h=h)
        c.border = border()

    # ── FEUILLE 1 : TABLEAU DE BORD ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tableau de Bord"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:H1")
    c = ws1["A1"]
    c.value = f"RAPPORT BANCAIRE — {banque} — {annee}"
    c.fill = fill("#0F172A")
    c.font = font(bold=True, size=16)
    c.alignment = align()
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells("A2:H2")
    c2 = ws1["A2"]
    c2.value = "Source : BCEAO · MongoDB Atlas · DataSphere Sénégal M2 Big Data ISM"
    c2.fill = fill("#1E3A5F")
    c2.font = font(size=9, color="B0C4DE")
    c2.alignment = align()
    ws1.row_dimensions[2].height = 20

    # KPIs
    row_annee = df_banque[df_banque["annee"] == annee]
    if row_annee.empty and not df_banque.empty:
        row_annee = df_banque.sort_values("annee").iloc[[-1]]

    def gv(col_n):
        if row_annee.empty or col_n not in row_annee.columns:
            return None
        v = row_annee[col_n].values[0]
        return float(v) if not pd.isna(v) else None

    kpi_items = [
        ("BILAN TOTAL",  gv("bilan"),                 "2563EB"),
        ("PNB",          gv("produit_net_bancaire"),  "059669"),
        ("RÉSULTAT NET", gv("resultat_net"),          "D97706"),
        ("FONDS PROPRES",gv("fonds_propres"),         "7C3AED"),
        ("EMPLOI",       gv("emploi"),                "0EA5E9"),
        ("RESSOURCES",   gv("ressources"),            "EC4899"),
        ("SOLVABILITÉ %",gv("ratio_solvabilite"),     "059669"),
        ("ROA %",        gv("ratio_rendement_actifs"),"D97706"),
    ]

    for i, (label, val, color_hex) in enumerate(kpi_items):
        col_start = (i % 4) * 2 + 1
        row_kpi = 4 + (i // 4) * 4

        ws1.merge_cells(start_row=row_kpi, start_column=col_start, end_row=row_kpi, end_column=col_start+1)
        c_lbl = ws1.cell(row=row_kpi, column=col_start, value=label)
        c_lbl.fill = fill(color_hex)
        c_lbl.font = font(bold=True, size=9)
        c_lbl.alignment = align()
        ws1.row_dimensions[row_kpi].height = 20

        ws1.merge_cells(start_row=row_kpi+1, start_column=col_start, end_row=row_kpi+2, end_column=col_start+1)
        if val is not None:
            if "%" in label:
                display = f"{val:.2f}%"
            elif abs(val) >= 1_000_000:
                display = f"{val/1_000_000:.2f}B FCFA"
            elif abs(val) >= 1_000:
                display = f"{val/1_000:.1f}Mds FCFA"
            else:
                display = f"{val:,.0f}M FCFA"
        else:
            display = "N/D"

        c_val = ws1.cell(row=row_kpi+1, column=col_start, value=display)
        c_val.fill = fill("EFF6FF")
        c_val.font = Font(name="Calibri", bold=True, size=13, color="0F172A")
        c_val.alignment = align()
        for r in range(row_kpi+1, row_kpi+3):
            ws1.row_dimensions[r].height = 18

    for i in range(1, 9):
        ws1.column_dimensions[get_column_letter(i)].width = 22

    # ── FEUILLE 2 : HISTORIQUE ────────────────────────────────────────
    ws2 = wb.create_sheet("Historique")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:J1")
    c = ws2["A1"]
    c.value = f"Évolution Historique — {banque}"
    c.fill = fill("#0F172A")
    c.font = font(bold=True, size=13)
    c.alignment = align()
    ws2.row_dimensions[1].height = 32

    cols_h = ["annee", "bilan", "produit_net_bancaire", "resultat_net",
              "fonds_propres", "emploi", "ressources",
              "ratio_solvabilite", "ratio_rendement_actifs", "ratio_rentabilite_capitaux"]
    hdrs_h = ["Année", "Bilan (M)", "PNB (M)", "Rés. Net (M)",
              "FP (M)", "Emploi (M)", "Ressources (M)", "Solvabilité %", "ROA %", "ROE %"]
    header_row(ws2, 2, hdrs_h, bg="1E3A5F")
    ws2.row_dimensions[2].height = 24

    for r_idx, (_, r) in enumerate(df_banque.iterrows(), 3):
        bg_c = "EFF6FF" if r_idx % 2 == 0 else "FFFFFF"
        for c_idx, col_n in enumerate(cols_h, 1):
            val = r.get(col_n)
            c_cell = ws2.cell(row=r_idx, column=c_idx)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                style_cell(c_cell, "N/D", bg=bg_c, fg="94A3B8")
            elif col_n == "annee":
                style_cell(c_cell, int(val), bg=bg_c, fg="0F172A", bold=True)
            elif col_n in ["ratio_solvabilite", "ratio_rendement_actifs", "ratio_rentabilite_capitaux"]:
                style_cell(c_cell, f"{float(val):.2f}%", bg=bg_c, fg="D97706", h="right")
            else:
                style_cell(c_cell, round(float(val), 0), bg=bg_c, fg="64748B", h="right")
        ws2.row_dimensions[r_idx].height = 18

    for i, w in enumerate([10,16,14,14,14,14,14,14,12,12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── FEUILLE 3 : RATIOS ───────────────────────────────────────────
    ws3 = wb.create_sheet("Ratios Financiers")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:F1")
    c = ws3["A1"]
    c.value = f"Analyse des Ratios Financiers — {banque}"
    c.fill = fill("#0F172A")
    c.font = font(bold=True, size=13)
    c.alignment = align()
    ws3.row_dimensions[1].height = 32

    ratio_items = [
        ("Solvabilité (FP/Bilan %)", "ratio_solvabilite",         "≥ 8%  (BCEAO)",  True,  8,  12),
        ("ROA (RN/Bilan %)",         "ratio_rendement_actifs",    "~1-2%",           True,  1,  2),
        ("ROE (RN/FP %)",            "ratio_rentabilite_capitaux","~10-15%",         True,  5,  10),
        ("Coeff. Exploit. (%)",      "coefficient_exploitation",  "<65%",            False, 65, 80),
        ("Liquidité (E/R %)",        "ratio_emplois_ressources",  "<100%",           False, 100, 110),
    ]
    header_row(ws3, 2, ["RATIO", "VALEUR", "NORME BCEAO", "STATUT", "INTERPRÉTATION"], bg="059669")
    ws3.row_dimensions[2].height = 24

    interp_map = {
        "ratio_solvabilite":         "Solidité financière réglementaire",
        "ratio_rendement_actifs":    "Efficacité à générer du profit sur les actifs",
        "ratio_rentabilite_capitaux":"Rendement pour les actionnaires",
        "coefficient_exploitation":  "Efficience opérationnelle (charges/revenu)",
        "ratio_emplois_ressources":  "Couverture des crédits par les dépôts",
    }

    for r_idx, (label, col_r, norme, plus_haut, s1, s2) in enumerate(ratio_items, 3):
        val_r = gv(col_r)
        bg_c = "EFF6FF" if r_idx % 2 == 0 else "FFFFFF"

        if val_r is not None and not (isinstance(val_r, float) and pd.isna(val_r)):
            val_txt = f"{val_r:.2f}%"
            if plus_haut:
                statut = "✅ Bon" if val_r >= s2 else ("⚠️ Moyen" if val_r >= s1 else "❌ Insuffisant")
                stat_col = "059669" if val_r >= s2 else ("D97706" if val_r >= s1 else "DC2626")
            else:
                statut = "✅ Bon" if val_r <= s1 else ("⚠️ Moyen" if val_r <= s2 else "❌ Élevé")
                stat_col = "059669" if val_r <= s1 else ("D97706" if val_r <= s2 else "DC2626")
        else:
            val_txt, statut, stat_col = "N/D", "—", "94A3B8"

        row_d = [label, val_txt, norme, statut, interp_map.get(col_r, "")]
        for c_idx, v in enumerate(row_d, 1):
            c_cell = ws3.cell(row=r_idx, column=c_idx)
            c_cell.value = v
            c_cell.fill = fill(bg_c)
            c_cell.font = Font(
                name="Calibri", size=10,
                color=stat_col if c_idx == 4 else "64748B",
                bold=(c_idx == 4),
            )
            c_cell.alignment = align(h="left" if c_idx in (1, 5) else "center")
            c_cell.border = border()
        ws3.row_dimensions[r_idx].height = 22

    for i, w in enumerate([28, 12, 15, 16, 32], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── FEUILLE 4 : POSITIONNEMENT ───────────────────────────────────
    ws4 = wb.create_sheet("Positionnement")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:E1")
    c = ws4["A1"]
    c.value = f"Positionnement Sectoriel — {annee}"
    c.fill = fill("#0F172A")
    c.font = font(bold=True, size=13)
    c.alignment = align()
    ws4.row_dimensions[1].height = 32

    if not df_all.empty and "bilan" in df_all.columns and "sigle" in df_all.columns:
        for cx in num_cols:
            if cx in df_all.columns:
                df_all[cx] = pd.to_numeric(df_all[cx], errors="coerce")
        classmt = (
            df_all[["sigle", "bilan", "fonds_propres", "produit_net_bancaire"]]
            .dropna(subset=["bilan"])
            .sort_values("bilan", ascending=False)
            .reset_index(drop=True)
        )
        classmt.index += 1
        header_row(ws4, 2, ["Rang", "Banque", "Bilan (M FCFA)", "FP (M FCFA)", "PNB (M FCFA)"], bg="D97706")
        ws4.row_dimensions[2].height = 24

        for r_idx, (idx, r2) in enumerate(classmt.iterrows(), 3):
            is_me = r2["sigle"] == banque
            bg_c = "E8F5E9" if is_me else ("EFF6FF" if r_idx % 2 == 0 else "FFFFFF")
            row_vals = [f"#{idx}", r2["sigle"],
                        round(float(r2.get("bilan") or 0), 0),
                        round(float(r2.get("fonds_propres") or 0), 0),
                        round(float(r2.get("produit_net_bancaire") or 0), 0)]
            for c_idx, val in enumerate(row_vals, 1):
                c_cell = ws4.cell(row=r_idx, column=c_idx)
                c_cell.value = val
                c_cell.fill = fill(bg_c)
                c_cell.font = Font(name="Calibri", size=10,
                                   color="059669" if is_me else "64748B", bold=is_me)
                c_cell.alignment = align(h="right" if c_idx > 2 else "center")
                c_cell.border = border()
            ws4.row_dimensions[r_idx].height = 18

        for i, w in enumerate([8, 16, 18, 16, 16], 1):
            ws4.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
